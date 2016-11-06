# Excel to Magento importer
from openpyxl import load_workbook
from magento import MagentoAPI
import math

# 1. get data from excel spreadsheet
wb = load_workbook(filename = 'spreadsheet.xlsx')
ws = wb['Sheet1']
cell_range = ws['B2':'B3']

# each row is a product configuration/variation or whatever you wanna call it (I will call it a variation)

# initialize some good old variables
old_title = ""

# a main array for everything
everything = {}
index = 0
# max rows 42422
for row in ws.iter_rows(min_row=2, max_row=42422):
	title = row[1].value
	if(old_title != title):
		if(old_title != ""):
			print("Saved to memory: " + old_title)
	
	# an array for this variation
	variation = {}

	if(old_title != title):
		# it's a fresh product so set the title
		old_title = title
		variation['is_first'] = 1
		variation['title'] = row[1].value
		variation['desc'] = row[2].value
		variation['paper_type'] = row[3].value
		variation['paper_weight'] = row[4].value
		variation['quantity'] = row[5].value
		variation['paper_size'] = row[6].value
		variation['sides'] = row[7].value
		variation['style'] = row[8].value
		variation['price'] = "{0:0.1f}".format(row[9].value)

		# gotta add the rest of the columns

	else:
		variation['is_first'] = 0
		variation['title'] = row[1].value
		variation['desc'] = row[2].value
		variation['paper_type'] = row[3].value
		variation['paper_weight'] = row[4].value
		variation['quantity'] = row[5].value
		variation['paper_size'] = row[6].value
		variation['sides'] = row[7].value
		variation['style'] = row[8].value
		variation['price'] = "{0:0.1f}".format(row[9].value)

		# gotta add the rest of the columns

	# add it to the everything array (because it's part of everything!)
	everything[index] = variation
	index += 1

# okay now all of the data is collected, let's move on the the good shit

# 2. connect to Magento

magento = MagentoAPI("magentohost.com", 80, "test_api_user", "test_api_key")

#to do:
# 3. import the products